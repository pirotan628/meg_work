import openpyxl
import pprint

def get_value_list(t_2d):
    return([[cell.value for cell in row] for row in t_2d])

def get_list_2d(sheet, start_row, end_row, start_col, end_col):
    return get_value_list(sheet.iter_rows(min_row=start_row,
                                          max_row=end_row,
                                          min_col=start_col,
                                          max_col=end_col))

wb = openpyxl.load_workbook('original_data.xlsx')

print(type(wb))
print(wb.sheetnames)

sheet = wb['後方視コホート2']

print(type(sheet))

#l_2d = get_list_2d(sheet, 1, 1895, 1, 196)
l_2d = get_list_2d(sheet, 1, 2, 1, 196)

pprint.pprint(l_2d, width=120)

#'通し','施設','症例番号','附数'
# 1,2,3,4
#脳波検査pre', '脳波検査post', '検査間隔(日)'
# 29,30,36
